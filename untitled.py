# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'untitled.ui'
#
# Created by: PyQt5 UI code generator 5.8
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtWidgets
import time

class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(640, 480)
        self.form = Form
        self.gridLayout = QtWidgets.QGridLayout(Form)
        self.gridLayout.setObjectName("gridLayout")

        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")

        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")

        self.label = QtWidgets.QLabel(Form)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)

        self.lineEdit = QtWidgets.QLineEdit(Form)
        self.lineEdit.setObjectName("lineEdit")
        self.horizontalLayout.addWidget(self.lineEdit)

        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout.addWidget(self.pushButton)

        self.verticalLayout.addLayout(self.horizontalLayout)

        self.tabWidget = QtWidgets.QTabWidget(Form)
        self.tabWidget.setObjectName("tabWidget")

        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")

        self.gridLayout_2 = QtWidgets.QGridLayout(self.tab)
        self.gridLayout_2.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_2.setObjectName("gridLayout_2")

        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")

        self.checkBox = QtWidgets.QCheckBox(self.tab)
        self.checkBox.setObjectName("checkBox")
        self.verticalLayout_2.addWidget(self.checkBox)

        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")

        self.label_2 = QtWidgets.QLabel(self.tab)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_2.addWidget(self.label_2)

        self.lineEdit_2 = QtWidgets.QLineEdit(self.tab)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.horizontalLayout_2.addWidget(self.lineEdit_2)

        self.verticalLayout_2.addLayout(self.horizontalLayout_2)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.listWidget = QtWidgets.QListWidget(self.tab)
        self.listWidget.setObjectName("listWidget")
        self.horizontalLayout_3.addWidget(self.listWidget)
        self.tableWidget = QtWidgets.QTableWidget(self.tab)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.horizontalLayout_3.addWidget(self.tableWidget)
        self.horizontalLayout_3.setStretch(0, 2)
        self.horizontalLayout_3.setStretch(1, 8)
        self.verticalLayout_2.addLayout(self.horizontalLayout_3)
        self.gridLayout_2.addLayout(self.verticalLayout_2, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.tab_2)
        self.gridLayout_3.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.checkBox_2 = QtWidgets.QCheckBox(self.tab_2)
        self.checkBox_2.setObjectName("checkBox_2")
        self.verticalLayout_3.addWidget(self.checkBox_2)
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.label_3 = QtWidgets.QLabel(self.tab_2)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout_4.addWidget(self.label_3)
        self.lineEdit_3 = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.horizontalLayout_4.addWidget(self.lineEdit_3)
        self.verticalLayout_3.addLayout(self.horizontalLayout_4)
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.listWidget_2 = QtWidgets.QListWidget(self.tab_2)
        self.listWidget_2.setObjectName("listWidget_2")
        self.horizontalLayout_5.addWidget(self.listWidget_2)
        self.tableWidget_2 = QtWidgets.QTableWidget(self.tab_2)
        self.tableWidget_2.setObjectName("tableWidget_2")
        self.tableWidget_2.setColumnCount(0)
        self.tableWidget_2.setRowCount(0)
        self.horizontalLayout_5.addWidget(self.tableWidget_2)
        self.horizontalLayout_5.setStretch(0, 2)
        self.horizontalLayout_5.setStretch(1, 8)
        self.verticalLayout_3.addLayout(self.horizontalLayout_5)
        self.gridLayout_3.addLayout(self.verticalLayout_3, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab_2, "")
        self.verticalLayout.addWidget(self.tabWidget)
        self.gridLayout.addLayout(self.verticalLayout, 0, 0, 1, 1)

        self.retranslateUi(Form)
        self._uiconnect()

        self.tabWidget.setCurrentIndex(0)
        # QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):

        Form.setWindowTitle("表格检查器")
        self.label.setText("文件路径")
        self.pushButton.setText("选择文件")
        self.checkBox.setText("仅显示错误信息")
        self.label_2.setText("消息")
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), "奖励表")
        self.checkBox_2.setText("仅显示错误信息")
        self.label_3.setText("消息")
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), "弟子表")

    def _uiconnect(self):
        self.pushButton.clicked.connect(self._msg)

    def _msg(self):

        filename, filetype = QtWidgets.QFileDialog.getOpenFileName(caption='打开', directory='c:\\',
                                                                   filter='EXCEL文件 (*.xlsx)')
        self.lineEdit.setText(filename)
        self.lineEdit.setEnabled(False)
        _sheetconf = self._xlsxhandle(filename=filename)
        self._showui(_sheetconf)

    def _xlsxhandle(self, filename):
        if filename == '':
            err_message = QtWidgets.QMessageBox.information(self.form, '提示', '选择的文件为空')
            return
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        sheetdict = {}
        sheetmaxrow = sheet.max_row
        sheetmaxcolumn = sheet.max_column

        for row in sheet.rows:
            for cell in row:
                sheetdict[(cell.row, cell.col_idx)] = str(cell.value)
        wb.close()
        _sheetconf = {'sheetmaxrow': sheetmaxrow, 'sheetmaxcolumn': sheetmaxcolumn, 'sheetdict': sheetdict}
        return _sheetconf  # 返回EXCEL最大行数、最大列数、表格内容

    def _showui(self, sheet):
        self.listWidget.clear()
        if sheet is not None:
            for i in range(sheet['sheetmaxrow']):
                self.listWidget.addItem(sheet['sheetdict'][(i + 1, 1)])


if __name__ == '__main__':
    import sys
    start=time.clock()
    app = QtWidgets.QApplication(sys.argv)
    qtobj = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(qtobj)
    qtobj.show()
    end=time.clock()
    print('time:%s'%(start-end))
    sys.exit(app.exec_())
