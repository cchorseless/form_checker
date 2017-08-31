import time

from openpyxl import load_workbook


class XlsxReader(object):
    def __init__(self):
        self.path = r'E:\fight__doc\execelTool\excel\item\item.xlsx'
        self.readxlsx()

    def readxlsx(self):
        wb = load_workbook(self.path,read_only=True)
        sheetlist = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(sheetlist[0])
        dictaa = []
        start = time.clock()
        for row in sheet.rows:
            for cell in row:
                dictaa.append(cell.value)
        # maxrow = sheet.max_row
        # maxcolumn = sheet.max_column
        # for i in range(maxrow):
        #     for j in range(maxcolumn):
        #         dictaa.append(sheet.cell(row=i + 1, column=j + 1).value)
        mid = time.clock()
        with open(r'C:\users\cchorseless\Desktop\111.txt', 'w+') as f:
            f.write(str(dictaa))
        end = time.clock()
        print('时间段%s时间段%s' % ((start - mid), (mid - end)))


if __name__ == '__main__':
    a = XlsxReader()
