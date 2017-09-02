# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'untitled88.ui'
#
# Created by: PyQt5 UI code generator 5.8
#
# WARNING! All changes made in this file will be lost!

import re
import sys
import time

from PyQt5 import QtCore, QtWidgets
from openpyxl import load_workbook

CHECK_STRING = ('0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '.', ';')


class Ui_Form(object):
    def setupUi(self, Form):  # 处理 widget的函数
        Form.setObjectName("Form")  # 对象名称
        Form.resize(640, 480)  # 表格尺寸
        self.Form = Form
        # 栅格布局
        self.gridLayout = QtWidgets.QGridLayout(Form)
        self.gridLayout.setObjectName("gridLayout")
        # 横向布局
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        # 标签
        self.label = QtWidgets.QLabel(Form)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        # 单行列表框
        self.lineEdit = QtWidgets.QLineEdit(Form)
        self.lineEdit.setObjectName("lineEdit")
        # self.lineEdit.setFocusPolicy(0)
        self.horizontalLayout.addWidget(self.lineEdit)
        # 按钮
        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout.addWidget(self.pushButton)
        # 将横向布局加入栅格布局中
        self.gridLayout.addLayout(self.horizontalLayout, 0, 0, 1, 1)
        # 数据表视图
        self.tabWidget = QtWidgets.QTabWidget(Form)
        self.tabWidget.setObjectName("tabWidget")
        #
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")

        self.gridLayout_2 = QtWidgets.QGridLayout(self.tab)
        self.gridLayout_2.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_2.setObjectName("gridLayout_2")

        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label_2 = QtWidgets.QLabel(self.tab)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_3.addWidget(self.label_2)

        self.lineEdit_2 = QtWidgets.QLineEdit(self.tab)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.horizontalLayout_3.addWidget(self.lineEdit_2)
        self.verticalLayout.addLayout(self.horizontalLayout_3)
        self.checkBox = QtWidgets.QCheckBox(self.tab)
        self.checkBox.setObjectName("checkBox")
        self.verticalLayout.addWidget(self.checkBox)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.listWidget = QtWidgets.QListWidget(self.tab)
        self.listWidget.setObjectName("listWidget")
        self.horizontalLayout_2.addWidget(self.listWidget)

        self.tableWidget = QtWidgets.QTableWidget(self.tab)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)

        self.horizontalLayout_2.addWidget(self.tableWidget)
        self.horizontalLayout_2.setStretch(0, 2)
        self.horizontalLayout_2.setStretch(1, 8)
        self.verticalLayout.addLayout(self.horizontalLayout_2)
        self.gridLayout_2.addLayout(self.verticalLayout, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab, "")

        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.tab_2)
        self.gridLayout_3.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.label_3 = QtWidgets.QLabel(self.tab_2)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout_4.addWidget(self.label_3)
        self.lineEdit_3 = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.horizontalLayout_4.addWidget(self.lineEdit_3)
        self.verticalLayout_2.addLayout(self.horizontalLayout_4)
        self.checkBox_2 = QtWidgets.QCheckBox(self.tab_2)
        self.checkBox_2.setObjectName("checkBox_2")
        self.verticalLayout_2.addWidget(self.checkBox_2)
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.listWidget_2 = QtWidgets.QListWidget(self.tab_2)
        self.listWidget_2.setObjectName("listWidget_2")
        self.horizontalLayout_5.addWidget(self.listWidget_2)

        self.tableWidget_2 = QtWidgets.QTableWidget(self.tab_2)
        self.tableWidget_2.setObjectName("tableWidget_2")
        self.tableWidget_2.setColumnCount(0)
        self.tableWidget_2.setRowCount(0)
        self.horizontalLayout_5.addWidget(self.tableWidget_2)
        self.horizontalLayout_5.setStretch(0, 2)
        self.horizontalLayout_5.setStretch(1, 8)
        self.verticalLayout_2.addLayout(self.horizontalLayout_5)
        self.gridLayout_3.addLayout(self.verticalLayout_2, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab_2, "")

        self.tab_3 = QtWidgets.QWidget()
        self.tab_3.setObjectName("tab_3")
        self.gridLayout_7 = QtWidgets.QGridLayout(self.tab_3)
        self.gridLayout_7.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_7.setObjectName("gridLayout_7")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_6.setObjectName("horizontalLayout_6")
        self.label_4 = QtWidgets.QLabel(self.tab_3)
        self.label_4.setObjectName("label_4")
        self.horizontalLayout_6.addWidget(self.label_4)
        self.lineEdit_4 = QtWidgets.QLineEdit(self.tab_3)
        self.lineEdit_4.setObjectName("lineEdit_4")
        self.horizontalLayout_6.addWidget(self.lineEdit_4)
        self.verticalLayout_3.addLayout(self.horizontalLayout_6)
        self.checkBox_3 = QtWidgets.QCheckBox(self.tab_3)
        self.checkBox_3.setObjectName("checkBox_3")
        self.verticalLayout_3.addWidget(self.checkBox_3)
        self.horizontalLayout_7 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_7.setObjectName("horizontalLayout_7")
        self.listWidget_3 = QtWidgets.QListWidget(self.tab_3)
        self.listWidget_3.setObjectName("listWidget_3")
        self.horizontalLayout_7.addWidget(self.listWidget_3)
        self.tableWidget_3 = QtWidgets.QTableWidget(self.tab_3)
        self.tableWidget_3.setObjectName("tableWidget_3")
        self.tableWidget_3.setColumnCount(0)
        self.tableWidget_3.setRowCount(0)

        self.horizontalLayout_7.addWidget(self.tableWidget_3)
        self.horizontalLayout_7.setStretch(0, 2)
        self.horizontalLayout_7.setStretch(1, 8)
        self.verticalLayout_3.addLayout(self.horizontalLayout_7)
        self.gridLayout_7.addLayout(self.verticalLayout_3, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab_3, "")
        self.tab_4 = QtWidgets.QWidget()
        self.tab_4.setObjectName("tab_4")
        self.gridLayout_6 = QtWidgets.QGridLayout(self.tab_4)
        self.gridLayout_6.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_6.setObjectName("gridLayout_6")
        self.gridLayout_4 = QtWidgets.QGridLayout()
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.gridLayout_5 = QtWidgets.QGridLayout()
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.label_5 = QtWidgets.QLabel(self.tab_4)
        self.label_5.setObjectName("label_5")
        self.gridLayout_5.addWidget(self.label_5, 0, 0, 1, 1)
        self.lineEdit_5 = QtWidgets.QLineEdit(self.tab_4)
        self.lineEdit_5.setObjectName("lineEdit_5")
        self.gridLayout_5.addWidget(self.lineEdit_5, 0, 1, 1, 1)
        self.gridLayout_4.addLayout(self.gridLayout_5, 0, 0, 1, 1)
        self.checkBox_4 = QtWidgets.QCheckBox(self.tab_4)
        self.checkBox_4.setObjectName("checkBox_4")
        self.gridLayout_4.addWidget(self.checkBox_4, 1, 0, 1, 1)
        self.horizontalLayout_9 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_9.setObjectName("horizontalLayout_9")
        self.listWidget_4 = QtWidgets.QListWidget(self.tab_4)
        self.listWidget_4.setObjectName("listWidget_4")
        self.horizontalLayout_9.addWidget(self.listWidget_4)
        self.tableWidget_4 = QtWidgets.QTableWidget(self.tab_4)
        self.tableWidget_4.setObjectName("tableWidget_4")
        self.tableWidget_4.setColumnCount(0)
        self.tableWidget_4.setRowCount(0)
        self.horizontalLayout_9.addWidget(self.tableWidget_4)
        self.horizontalLayout_9.setStretch(0, 2)
        self.horizontalLayout_9.setStretch(1, 8)
        self.gridLayout_4.addLayout(self.horizontalLayout_9, 2, 0, 1, 1)
        self.gridLayout_6.addLayout(self.gridLayout_4, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab_4, "")
        self.gridLayout.addWidget(self.tabWidget, 1, 0, 1, 1)
        self.retranslateUi(Form)
        self.UiLogic()
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "表格检查器"))
        self.label.setText(_translate("Form", "文件"))
        self.pushButton.setText(_translate("Form", "打开文件"))
        self.label_2.setText(_translate("Form", "消息"))
        self.checkBox.setText(_translate("Form", "筛选错误信息"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("Form", "奖励表"))
        self.label_3.setText(_translate("Form", "消息"))
        self.checkBox_2.setText(_translate("Form", "筛选错误信息"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("Form", "物品表"))
        self.label_4.setText(_translate("Form", "消息"))
        self.checkBox_3.setText(_translate("Form", "筛选错误信息"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_3), _translate("Form", "弟子表"))
        self.label_5.setText(_translate("Form", "消息"))
        self.checkBox_4.setText(_translate("Form", "筛选错误信息"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_4), _translate("Form", "弟子表"))

    def UiLogic(self):
        self.pushButton.clicked.connect(self.msg)

    def msg(self):
        filename, filetype = QtWidgets.QFileDialog.getOpenFileName(caption='打开',
                                                                   directory=r'E:\fight__doc\execelTool\excel\item',
                                                                   filter='Excel文件 (*.xlsx)')
        self.lineEdit.setText(filename)
        self.lineEdit.setEnabled(False)

        wb = load_workbook(filename)
        sheetlist = wb.get_sheet_names()
        self.sheet = wb.get_sheet_by_name(sheetlist[0])
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setHorizontalHeaderLabels(['ID', '名称', '稀有度', '概率', '数量', '装备相关'])
        # self.tableWidget.setRowCount(self.sheet.max_row)
        self.sheetdict = {}
        start = time.clock()
        for row in self.sheet.rows:
            for cell in row:
                self.sheetdict[(cell.row, cell.col_idx)] = str(cell.value)
        print(self.sheetdict)
        mid = time.clock()
        self.work()
        #
        # 多线程
        # threadinglist = []
        # for i in range(1):
        #     t = threading.Thread(target=self.work, name=i)
        #     threadinglist.append(t)
        # for t in threadinglist:
        #     t.start()
        # for t in threadinglist:
        #     t.join()
        self.prize_logic()
        end = time.clock()
        print('所需时间：%s ; 所需时间：%s ' % ((start - mid), (mid - end)))

    def work(self):

        for i in range(3, self.sheet.max_row):
            item = '%s--%s' % (self.sheetdict[(i + 1, 1)], self.sheetdict[(i + 1, 2)])
            self.listWidget.addItem(item)

        self.listWidget.itemClicked.connect(self.aaaa)

    def aaaa(self, obj):
        listindex = self.listWidget.row(obj)
        aaadict = self.prizerowlist[listindex]

        # 奖励表
        idlist = [x for x in aaadict['item'] if x != 0]
        numberlist = [x for x in aaadict['number'] if x != 0]
        quanzhonglist = [x for x in aaadict['quanzhong'] if x != 0]
        equipranddictlist = [x for x in aaadict['equipranddict'] if x != 0]
        # itemlist=[x for x in aaadict['item'] if x !=0]
        # itemlist=[x for x in aaadict['item'] if x !=0]
        # itemlist=[x for x in aaadict['item'] if x !=0]

        # print(idlist)
        columnindex = len(idlist)
        self.tableWidget.setRowCount(columnindex)
        if columnindex != 0:
            for i in range(columnindex):
                # 物品ID,第一列
                item_1 = QtWidgets.QTableWidgetItem(str(idlist[i]))
                self.tableWidget.setItem(i, 0, item_1)
                # 物品名称,第二列
                # item_2 = QtWidgets.QTableWidgetItem(str(numberlist[i]))
                # self.tableWidget.setItem(i, 1, item_2)
                # 物品稀有度,第三列
                # item_3 = QtWidgets.QTableWidgetItem(str(aaalist[i]))
                # self.tableWidget.setItem(i, 2, item_3)
                # 物品概率,第四列
                item_4 = QtWidgets.QTableWidgetItem(str(quanzhonglist[i]))
                self.tableWidget.setItem(i, 3, item_4)
                # 物品数量，第五列
                item_5 = QtWidgets.QTableWidgetItem(str(numberlist[i]))
                self.tableWidget.setItem(i, 4, item_5)
                # 装备相关，第六列
                # item_6 = QtWidgets.QTableWidgetItem(str(equipranddictlist[i]))
                # self.tableWidget.setItem(i, 5, item_6)

    def prize_logic(self):

        gamemoney, gameequip, gameitem, gamepanter, prizerow = {}, {}, {}, {}, {}
        gamemoney_len, gameequip_len, gameitem_len, gamepanter_len, prizerow_len = {}, {}, {}, {}, {}
        self.prizerowlist = []
        prizesheet = self.sheetdict
        error_list = []

        for i in range(4, self.sheet.max_row + 1):
            gamemoney[i] = [prizesheet[(i, j)].split(';') for j in range(4, 7)]
            gameequip[i] = [prizesheet[(i, j)].split(';') for j in range(7, 11)]
            gameitem[i] = [prizesheet[(i, j)].split(';') for j in range(11, 14)]
            gamepanter[i] = [prizesheet[(i, j)].split(';') for j in range(14, 17)]
            prizerow[i] = [gamemoney[i], gameequip[i], gameitem[i], gamepanter[i]]

            gamemoney_len[i] = [len(re.split('\;|\,|\:|\；|\ ', prizesheet[(i, j)])) - re.split('\;|\,|\:|\；|\ ',
                                                                                               prizesheet[
                                                                                                   (i, j)]).count(
                'None') for
                                j
                                in range(4, 7)]
            gameequip_len[i] = [len(re.split('\;|\,|\:|\；|\ ', prizesheet[(i, j)])) - re.split('\;|\,|\:|\；|\ ',
                                                                                               prizesheet[
                                                                                                   (i, j)]).count(
                'None') for j
                                in range(7, 11)]
            gameitem_len[i] = [len(re.split('\;|\,|\:|\；|\ ', prizesheet[(i, j)])) - re.split('\;|\,|\:|\；|\ ',
                                                                                              prizesheet[
                                                                                                  (i, j)]).count(
                'None') for j in
                               range(11, 14)]
            gamepanter_len[i] = [len(re.split('\;|\,|\:|\；|\ ', prizesheet[(i, j)])) - re.split('\;|\,|\:|\；|\ ',
                                                                                                prizesheet[
                                                                                                    (i, j)]).count(
                'None') for j
                                 in range(14, 17)]

            prizerow_len[i] = [gamemoney_len[i], gameequip_len[i], gameitem_len[i], gamepanter_len[i]]

            for k in range(4):
                # 数据检查，添加错误列表
                if prizerow_len[i][k].count(prizerow_len[i][k][0]) != len(prizerow_len[i][k]):
                    error_list.append((i, k))

            # 转化概率
            prizerowdict = {}
            prizerowdict['quanzhong'] = list(map(eval,
                                                 prizerow[i][0][2] + prizerow[i][1][3] + prizerow[i][2][2] +
                                                 prizerow[i][3][
                                                     2]))

            prizerowdict['item'] = list(map(eval,
                                            prizerow[i][0][0] + prizerow[i][1][0] + prizerow[i][2][0] + prizerow[i][3][
                                                0]))
            prizerowdict['number'] = list(map(eval,
                                              prizerow[i][0][1] + prizerow[i][1][1] + prizerow[i][2][1] +
                                              prizerow[i][3][1]))
            prizerowdict['equipranddict'] = list(map(eval, prizerow[i][1][2]))
            prizerowdict['id'] = [i - 4]
            list_linshi = list(prizerowdict.values())

            for i in range(5):
                for index, value in enumerate(list_linshi[i]):
                    if value == None:
                        list_linshi[i][index] = 0
            print(prizerowdict['quanzhong'])
            print(sum(prizerowdict['quanzhong']))
            for index_1, value_1 in enumerate(prizerowdict['quanzhong']):
                if sum(prizerowdict['quanzhong']) == 0:
                    prizerowdict['quanzhong'][index_1] = 0
                else:
                    prizerowdict['quanzhong'][index_1] = value_1 / sum(prizerowdict['quanzhong'])
            print(sum(prizerowdict['quanzhong']))
            self.prizerowlist.append(prizerowdict)

        print(prizerow)
        print(prizerow_len)
        print(error_list)
        # print(prizerowdict)
        print(self.prizerowlist)


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)  # 注册UI界面
    aaa = QtWidgets.QWidget()  # 实例化UI界面
    ui = Ui_Form()  # 实例化对象
    ui.setupUi(aaa)  # 对象处理UI界面
    aaa.show()
    sys.exit(app.exec_())  # 保持界面不退出，接受指令退出
