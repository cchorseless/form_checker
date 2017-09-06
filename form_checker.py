# -*- coding: utf-8 -*-

import os
import re
import sys

# Form implementation generated from reading ui file 'form_checker.ui'
#
# Created by: PyQt5 UI code generator 5.8
#
# WARNING! All changes made in this file will be lost!
from PyQt5 import QtCore, QtWidgets, QtGui
from openpyxl import load_workbook


class Ui_Form(object):
    def setupUi(self, Form):
        # 低层表格对象
        Form.setObjectName("Form")
        Form.resize(833, 520)
        self.form = Form
        #
        self.gridLayout = QtWidgets.QGridLayout(Form)
        self.gridLayout.setObjectName("gridLayout")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem)
        self.label = QtWidgets.QLabel(Form)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        self.lineEdit = QtWidgets.QLineEdit(Form)
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit.setEnabled(False)
        self.horizontalLayout.addWidget(self.lineEdit)
        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout.addWidget(self.pushButton)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem1)
        self.horizontalLayout.setStretch(0, 1)
        self.horizontalLayout.setStretch(1, 1)
        self.horizontalLayout.setStretch(2, 3)
        self.horizontalLayout.setStretch(3, 1)
        self.horizontalLayout.setStretch(4, 1)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.line = QtWidgets.QFrame(Form)
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.verticalLayout.addWidget(self.line)
        self.tabWidget = QtWidgets.QTabWidget(Form)
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.tab)
        self.gridLayout_2.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.label_2 = QtWidgets.QLabel(self.tab)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_2.addWidget(self.label_2)
        self.progressBar = QtWidgets.QProgressBar(self.tab)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.horizontalLayout_2.addWidget(self.progressBar)
        self.pushButton_2 = QtWidgets.QPushButton(self.tab)
        self.pushButton_2.setObjectName("pushButton_2")
        self.horizontalLayout_2.addWidget(self.pushButton_2)
        self.verticalLayout_3.addLayout(self.horizontalLayout_2)
        self.line_2 = QtWidgets.QFrame(self.tab)
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.verticalLayout_3.addWidget(self.line_2)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.checkBox = QtWidgets.QCheckBox(self.tab)
        self.checkBox.setObjectName("checkBox")
        self.verticalLayout_2.addWidget(self.checkBox)
        self.listWidget = QtWidgets.QListWidget(self.tab)
        self.listWidget.setObjectName("listWidget")
        self.verticalLayout_2.addWidget(self.listWidget)
        self.horizontalLayout_3.addLayout(self.verticalLayout_2)
        self.tableWidget = QtWidgets.QTableWidget(self.tab)
        self.tableWidget.setObjectName("tableWidget")
        self.horizontalLayout_3.addWidget(self.tableWidget)
        self.horizontalLayout_3.setStretch(0, 2)
        self.horizontalLayout_3.setStretch(1, 8)
        self.verticalLayout_3.addLayout(self.horizontalLayout_3)
        self.gridLayout_2.addLayout(self.verticalLayout_3, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.tab_2)
        self.gridLayout_3.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.verticalLayout_4 = QtWidgets.QVBoxLayout()
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.label_3 = QtWidgets.QLabel(self.tab_2)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout_4.addWidget(self.label_3)
        self.progressBar_2 = QtWidgets.QProgressBar(self.tab_2)
        self.progressBar_2.setProperty("value", 0)
        self.progressBar_2.setObjectName("progressBar_2")
        self.horizontalLayout_4.addWidget(self.progressBar_2)
        self.pushButton_3 = QtWidgets.QPushButton(self.tab_2)
        self.pushButton_3.setObjectName("pushButton_3")
        self.horizontalLayout_4.addWidget(self.pushButton_3)
        self.verticalLayout_4.addLayout(self.horizontalLayout_4)
        self.line_3 = QtWidgets.QFrame(self.tab_2)
        self.line_3.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_3.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_3.setObjectName("line_3")
        self.verticalLayout_4.addWidget(self.line_3)
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout()
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.checkBox_2 = QtWidgets.QCheckBox(self.tab_2)
        self.checkBox_2.setObjectName("checkBox_2")
        self.verticalLayout_5.addWidget(self.checkBox_2)
        self.listWidget_2 = QtWidgets.QListWidget(self.tab_2)
        self.listWidget_2.setObjectName("listWidget_2")
        self.verticalLayout_5.addWidget(self.listWidget_2)
        self.horizontalLayout_5.addLayout(self.verticalLayout_5)
        self.tableWidget_2 = QtWidgets.QTableWidget(self.tab_2)
        self.tableWidget_2.setObjectName("tableWidget_2")
        self.horizontalLayout_5.addWidget(self.tableWidget_2)
        self.horizontalLayout_5.setStretch(0, 2)
        self.horizontalLayout_5.setStretch(1, 8)
        self.verticalLayout_4.addLayout(self.horizontalLayout_5)
        self.gridLayout_3.addLayout(self.verticalLayout_4, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab_2, "")
        self.tab_3 = QtWidgets.QWidget()
        self.tab_3.setObjectName("tab_3")
        self.gridLayout_4 = QtWidgets.QGridLayout(self.tab_3)
        self.gridLayout_4.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.verticalLayout_6 = QtWidgets.QVBoxLayout()
        self.verticalLayout_6.setObjectName("verticalLayout_6")
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_6.setObjectName("horizontalLayout_6")
        self.label_4 = QtWidgets.QLabel(self.tab_3)
        self.label_4.setObjectName("label_4")
        self.horizontalLayout_6.addWidget(self.label_4)
        self.progressBar_3 = QtWidgets.QProgressBar(self.tab_3)
        self.progressBar_3.setProperty("value", 0)
        self.progressBar_3.setObjectName("progressBar_3")
        self.horizontalLayout_6.addWidget(self.progressBar_3)
        self.pushButton_4 = QtWidgets.QPushButton(self.tab_3)
        self.pushButton_4.setObjectName("pushButton_4")
        self.horizontalLayout_6.addWidget(self.pushButton_4)
        self.verticalLayout_6.addLayout(self.horizontalLayout_6)
        self.line_4 = QtWidgets.QFrame(self.tab_3)
        self.line_4.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_4.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_4.setObjectName("line_4")
        self.verticalLayout_6.addWidget(self.line_4)
        self.horizontalLayout_7 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_7.setObjectName("horizontalLayout_7")
        self.verticalLayout_7 = QtWidgets.QVBoxLayout()
        self.verticalLayout_7.setObjectName("verticalLayout_7")
        self.checkBox_3 = QtWidgets.QCheckBox(self.tab_3)
        self.checkBox_3.setObjectName("checkBox_3")
        self.verticalLayout_7.addWidget(self.checkBox_3)
        self.listWidget_3 = QtWidgets.QListWidget(self.tab_3)
        self.listWidget_3.setObjectName("listWidget_3")
        self.verticalLayout_7.addWidget(self.listWidget_3)
        self.horizontalLayout_7.addLayout(self.verticalLayout_7)
        self.tableWidget_3 = QtWidgets.QTableWidget(self.tab_3)
        self.tableWidget_3.setObjectName("tableWidget_3")
        self.horizontalLayout_7.addWidget(self.tableWidget_3)
        self.horizontalLayout_7.setStretch(0, 2)
        self.horizontalLayout_7.setStretch(1, 8)
        self.verticalLayout_6.addLayout(self.horizontalLayout_7)
        self.gridLayout_4.addLayout(self.verticalLayout_6, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab_3, "")
        self.tab_4 = QtWidgets.QWidget()
        self.tab_4.setObjectName("tab_4")
        self.gridLayout_7 = QtWidgets.QGridLayout(self.tab_4)
        self.gridLayout_7.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_7.setObjectName("gridLayout_7")
        self.gridLayout_5 = QtWidgets.QGridLayout()
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.gridLayout_6 = QtWidgets.QGridLayout()
        self.gridLayout_6.setObjectName("gridLayout_6")
        self.label_5 = QtWidgets.QLabel(self.tab_4)
        self.label_5.setObjectName("label_5")
        self.gridLayout_6.addWidget(self.label_5, 0, 0, 1, 1)
        self.progressBar_4 = QtWidgets.QProgressBar(self.tab_4)
        self.progressBar_4.setProperty("value", 0)
        self.progressBar_4.setObjectName("progressBar_4")
        self.gridLayout_6.addWidget(self.progressBar_4, 0, 1, 1, 1)
        self.pushButton_5 = QtWidgets.QPushButton(self.tab_4)
        self.pushButton_5.setObjectName("pushButton_5")
        self.gridLayout_6.addWidget(self.pushButton_5, 0, 2, 1, 1)
        self.gridLayout_5.addLayout(self.gridLayout_6, 0, 0, 1, 1)
        self.line_5 = QtWidgets.QFrame(self.tab_4)
        self.line_5.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_5.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_5.setObjectName("line_5")
        self.gridLayout_5.addWidget(self.line_5, 1, 0, 1, 1)
        self.horizontalLayout_9 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_9.setObjectName("horizontalLayout_9")
        self.verticalLayout_9 = QtWidgets.QVBoxLayout()
        self.verticalLayout_9.setObjectName("verticalLayout_9")
        self.checkBox_4 = QtWidgets.QCheckBox(self.tab_4)
        self.checkBox_4.setObjectName("checkBox_4")
        self.verticalLayout_9.addWidget(self.checkBox_4)
        self.listWidget_4 = QtWidgets.QListWidget(self.tab_4)
        self.listWidget_4.setObjectName("listWidget_4")
        self.verticalLayout_9.addWidget(self.listWidget_4)
        self.horizontalLayout_9.addLayout(self.verticalLayout_9)
        self.tableWidget_4 = QtWidgets.QTableWidget(self.tab_4)
        self.tableWidget_4.setObjectName("tableWidget_4")
        self.horizontalLayout_9.addWidget(self.tableWidget_4)
        self.horizontalLayout_9.setStretch(0, 2)
        self.horizontalLayout_9.setStretch(1, 8)
        self.gridLayout_5.addLayout(self.horizontalLayout_9, 2, 0, 1, 1)
        self.gridLayout_7.addLayout(self.gridLayout_5, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab_4, "")
        self.verticalLayout.addWidget(self.tabWidget)
        self.gridLayout.addLayout(self.verticalLayout, 0, 0, 1, 1)

        self.retranslateUi(Form)
        self._uiconnect()
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "表格检查器"))
        self.label.setText(_translate("Form", "选择文件路径"))
        self.pushButton.setText(_translate("Form", "打开文件"))
        self.label_2.setText(_translate("Form", "解析进度"))
        self.pushButton_2.setText(_translate("Form", "开始解析"))
        self.checkBox.setText(_translate("Form", "筛选显示错误"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("Form", "奖励表"))
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setHorizontalHeaderLabels(['ID', '名称', '稀有度', '概率', '数量', '装备相关'])
        self.label_3.setText(_translate("Form", "解析进度"))
        self.pushButton_3.setText(_translate("Form", "开始解析"))
        self.checkBox_2.setText(_translate("Form", "筛选显示错误"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("Form", "弟子表"))
        self.label_4.setText(_translate("Form", "解析进度"))
        self.pushButton_4.setText(_translate("Form", "开始解析"))
        self.checkBox_3.setText(_translate("Form", "筛选显示错误"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_3), _translate("Form", "关卡表"))
        self.label_5.setText(_translate("Form", "解析进度"))
        self.pushButton_5.setText(_translate("Form", "开始解析"))
        self.checkBox_4.setText(_translate("Form", "筛选显示错误"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_4), _translate("Form", "天赋表"))

    # 界面连接配置
    def _uiconnect(self):
        self.pushButton.clicked.connect(self._msg)
        self.pushButton_2.clicked.connect(self._prize_handle)
        self.pushButton_3.clicked.connect(self._partnerhandle)
        self.pushButton_4.clicked.connect(self._missionhandle)
        self.pushButton_5.clicked.connect(self._geniushandle)
        self.listWidget.itemClicked.connect(self._prize_listitem_click)

    # 选取文件夹，匹配文件
    def _msg(self):
        self.filespathdict = {}
        self.lineEdit.clear()
        exceldirpath = QtWidgets.QFileDialog.getExistingDirectory(self.form, caption='请选择EXCEL文件夹', directory='e:\\')
        self.lineEdit.setText(exceldirpath)
        if exceldirpath == '':
            # 弹出警告框
            errorbox = QtWidgets.QMessageBox.information(self.form, '提示', '未选择任何文件')
            return
        else:
            for root, dirs, files in os.walk(exceldirpath):
                for f in files:
                    if f in ['item.xlsx', 'prize.xlsx', 'mission.xlsx', 'partner.xlsx', 'money.xlsx',
                             'genius.xlsx', 'equip.xlsx', 'equip_random_attr.xlsx']:
                        self.filespathdict[f] = os.path.join(root, f)

    # 读取表格内容
    def _xlshandle(self, filename):
        filepath = self.filespathdict[filename]
        wb = load_workbook(filepath)
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        sheetdict = {}
        sheetmaxrow = sheet.max_row
        sheetmaxcolumn = sheet.max_column
        for row in sheet.rows:
            for cell in row:
                sheetdict[(cell.row, cell.col_idx)] = str(cell.value)
        wb.close()
        _sheetconf = {'sheetmaxrow': sheetmaxrow, 'sheetmaxcolumn': sheetmaxcolumn, 'sheetdict': sheetdict}
        return _sheetconf

    def _prize_handle(self):

        gamemoney, gameequip, gameitem, gamepanter, prizerow = {}, {}, {}, {}, {}
        self.listWidget.clear()
        sheetconf = self._xlshandle('prize.xlsx')
        prizesheet = sheetconf['sheetdict']
        prizerowlist = []
        error_list = []

        for i in range(4, sheetconf['sheetmaxrow'] + 1):
            prizerowdict = {}  # 每次循环初始化这个字典容器
            errstate = False
            # 数据检查，添加错误列表—————————————————————————————————
            for j in range(4, 17):
                if prizesheet[(i, j)] != 'None':
                    for item in list(prizesheet[(i, j)]):
                        tmp = re.findall('[^\d;]', item)
                        if tmp:
                            error_list.append((i, j))
                            errstate = True

            gamemoney[i] = [prizesheet[(i, j)].split(';') for j in range(4, 7)]
            gameequip[i] = [prizesheet[(i, j)].split(';') for j in range(7, 11)]
            gameitem[i] = [prizesheet[(i, j)].split(';') for j in range(11, 14)]
            gamepanter[i] = [prizesheet[(i, j)].split(';') for j in range(14, 17)]
            prizerow[i] = [gamemoney[i], gameequip[i], gameitem[i], gamepanter[i]]
            for itemdict in prizerow[i]:
                item_len = []
                for itemlist in itemdict:
                    if itemlist == ['None']:
                        item_len.append(0)
                    else:
                        item_len.append(len(itemlist))
                if item_len.count(item_len[0]) != len(item_len):
                    error_list.append((i, prizerow[i].index(itemdict)))
                    errstate = True

            try:
                prizerowdict['quanzhong'] = list(map(eval,
                                                     prizerow[i][0][2] + prizerow[i][1][3] + prizerow[i][2][2] +
                                                     prizerow[i][3][2]))
            except Exception:
                error_list.append((i, 'quanzhong'))
                errstate = True
            try:
                prizerowdict['item'] = list(map(eval,
                                                prizerow[i][0][0] + prizerow[i][1][0] + prizerow[i][2][0] +
                                                prizerow[i][3][
                                                    0]))
            except Exception:
                error_list.append((i, 'item'))
                errstate = True
            try:
                prizerowdict['number'] = list(map(eval,
                                                  prizerow[i][0][1] + prizerow[i][1][1] + prizerow[i][2][1] +
                                                  prizerow[i][3][1]))
            except Exception:
                error_list.append((i, 'number'))
                errstate = True
            try:
                prizerowdict['equipranddict'] = list(map(eval, prizerow[i][1][2]))
            except Exception:
                error_list.append((i, 'equipranddict'))
                errstate = True

            if errstate == True:
                prizerowdict = {'number': [0], 'equipranddict': [0], 'quanzhong': [0], 'item': [0]}

            prizerowdict['id'] = [i - 4]
            # 转化概率----------------------------------------------------------------------------
            # 将prizerowdict中None替换为0
            for key in prizerowdict.keys():
                for idx, value in enumerate(prizerowdict[key]):
                    if value is None:
                        prizerowdict[key][idx] = 0
                        ##############################################
            sumprizequanzhong = sum(prizerowdict['quanzhong'])
            for index, value in enumerate(prizerowdict['quanzhong']):
                if prizerowdict['quanzhong'][index] != 0:
                    prizerowdict['quanzhong'][index] = value / sumprizequanzhong
            prizerowlist.append(prizerowdict)

        # ############################转化概率结束########
        # 标识错误条目—————————————————
        if sheetconf != {}:
            for i in range(4, sheetconf['sheetmaxrow'] + 1):
                self.listWidget.addItem(sheetconf['sheetdict'][(i, 1)])
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 0))
        brush.setStyle(QtCore.Qt.Dense1Pattern)

        if error_list != []:
            for tmp in error_list:
                self.listWidget.item(tmp[0] - 4).setBackground(brush)
        # -----------------------------------------------------
        self._prize_error_list = error_list
        self._prize_prizerowlist = prizerowlist

    def _prize_listitem_click(self, obj):

        listindex = self.listWidget.row(obj)
        print(listindex)
        print(self._prize_error_list)

        for item in self._prize_error_list:
            if listindex == item[0] - 4:
                self.tableWidget.setRowCount(1)
                erritem = QtWidgets.QTableWidgetItem('此条数据报错，请修改')
                self.tableWidget.setItem(0, 0, erritem)
            else:
                aaadict = self._prize_prizerowlist[listindex]
                idlist = [x for x in aaadict['item'] if x != 0]
                numberlist = [x for x in aaadict['number'] if x != 0]
                quanzhonglist = [x for x in aaadict['quanzhong'] if x != 0]
                equipranddictlist = [x for x in aaadict['equipranddict'] if x != 0]
                # itemlist=[x for x in aaadict['item'] if x !=0]
                # itemlist=[x for x in aaadict['item'] if x !=0]
                # itemlist=[x for x in aaadict['item'] if x !=0]
                columnindex = len(idlist)
                self.tableWidget.setRowCount(columnindex)
                if columnindex != 0:
                    for i in range(columnindex):
                        # 物品ID,第一列
                        item_1 = QtWidgets.QTableWidgetItem(str(idlist[i]))
                        self.tableWidget.setItem(i, 0, item_1)
                        # 物品名称,第二列
                        # item_2 = QtWidgets.QTableWidgetItem(str(numberlist[i]))
                        # self.tableWidget.setItem(i, 1, item_2)
                        # 物品稀有度,第三列
                        # item_3 = QtWidgets.QTableWidgetItem(str(aaalist[i]))
                        # self.tableWidget.setItem(i, 2, item_3)
                        # 物品概率,第四列
                        item_4 = QtWidgets.QTableWidgetItem(str(quanzhonglist[i]))
                        self.tableWidget.setItem(i, 3, item_4)
                        # 物品数量，第五列
                        item_5 = QtWidgets.QTableWidgetItem(str(numberlist[i]))
                        self.tableWidget.setItem(i, 4, item_5)
                        # 装备相关，第六列
                        # item_6 = QtWidgets.QTableWidgetItem(str(equipranddictlist[i]))
                        # self.tableWidget.setItem(i, 5, item_6)

    def _partnerhandle(self):
        pass

    def _missionhandle(self):
        pass

    def _geniushandle(self):
        pass


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    obj = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(obj)
    obj.show()
    sys.exit(app.exec_())
